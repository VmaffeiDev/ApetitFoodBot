"""Leitura da planilha do Excel para as mesmas linhas que o CSV produz.

A operacao manda .xlsx tanto quanto .csv. Em vez de um segundo importador,
esta camada so converte a planilha em linhas de texto e devolve para o
`csv_import`, que continua sendo o unico lugar onde regra de layout mora.

Converter para texto de proposito: no CSV tudo chega como string, e fazer o
Excel entregar float aqui abriria um caminho em que o mesmo arquivo importa
diferente por ter vindo em outro formato.
"""

from pathlib import Path

SPREADSHEET_SUFFIXES = {".xlsx", ".xlsm", ".xltx", ".xltm"}


def is_spreadsheet(path: str | Path) -> bool:
    return Path(path).suffix.lower() in SPREADSHEET_SUFFIXES


def _cell_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def read_spreadsheet_rows(path, sheet: str = "") -> list[list[str]]:
    """Devolve as linhas da planilha como texto, no formato do `csv_import`.

    Aceita caminho no disco ou objeto de arquivo ja aberto em memoria — o
    arquivo que chega pelo Telegram nunca toca o disco.

    Sem `sheet`, le a primeira aba — que e onde os exports observados poem o
    cardapio.
    """
    try:
        import openpyxl
    except ImportError as erro:  # pragma: no cover - depende do ambiente
        raise RuntimeError(
            "Ler .xlsx precisa do openpyxl. Instale com: pip install openpyxl — "
            "ou salve a planilha como CSV e importe o CSV."
        ) from erro

    workbook = openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        if sheet:
            if sheet not in workbook.sheetnames:
                disponiveis = ", ".join(workbook.sheetnames)
                raise ValueError(f'Aba "{sheet}" nao existe. Abas do arquivo: {disponiveis}.')
            worksheet = workbook[sheet]
        else:
            worksheet = workbook.worksheets[0]
        return [[_cell_text(cell) for cell in row] for row in worksheet.iter_rows(values_only=True)]
    finally:
        workbook.close()
